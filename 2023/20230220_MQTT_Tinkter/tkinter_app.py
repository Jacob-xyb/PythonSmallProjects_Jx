import tkinter as tk
import tkinter_tool
from tkinter import ttk
import py_tool
from MQTT.test import *
from tkinter.filedialog import askopenfilename
import openpyxl
from DealCase import *
from LogTool import *
import threading
import os
from collections import defaultdict


# https://sr03qc8xc3.feishu.cn/docx/TENEd7b9Boh6XaxrXbpcXodmn26


ConfigDict = {}
if os.path.exists("config.inf"):
    ConfigDict = py_tool.load("config.inf")

mqtt_client = MqttRoad()
ssh_client = SSHRoad()


PyLog().init_logDIr()


def thread_it(func, *args):
    '''将函数打包进线程'''
    # 创建
    t = threading.Thread(target=func, args=args)
    # 守护 !!!
    t.setDaemon(True)
    # 启动
    t.start()
    # 阻塞--卡死界面！
    # t.join()


class tk_ui:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title('MQTT')
        self.init_config()
        self.left_frame = tk.Frame(self.window)
        self.center_frame = tk.Frame(self.window)
        self.right_frame = tk.Frame(self.window)
        self.bottom_frame = tk.Frame(self.window)

        self.left_frame.grid(row=0, column=0, sticky="n", padx=4)
        self.center_frame.grid(row=0, column=1, sticky="n", padx=4)
        self.right_frame.grid(row=0, column=2, sticky="n", padx=4, rowspan=2)
        self.bottom_frame.grid(row=1, column=0, sticky="w", padx=4, columnspan=2)

        self.set_mqtt_connect()
        self.set_ssh_connect()
        self.set_subscribe()
        self.set_message_validate()
        self.set_message_result()
        self.set_case_table()
        self.set_case_result_table()
        self.window.protocol('WM_DELETE_WINDOW', self.QuitWindow)

    def QuitWindow(self):
        ConfigDict["mqtt_url_text"] = self.mqtt_url_text.get()
        ConfigDict["ssh_url_text"] = self.ssh_url_text.get()
        ConfigDict["ssh_username_text"] = self.ssh_username_text.get()
        ConfigDict["ssh_password_text"] = self.ssh_password_text.get()
        py_tool.save(ConfigDict, "config.inf")
        self.window.destroy()

    def init_config(self):
        self.mqtt_url_text = tk.StringVar()
        self.mqtt_url_text.set(ConfigDict.get("mqtt_url_text", ""))
        self.ssh_url_text = tk.StringVar()
        self.ssh_url_text.set(ConfigDict.get("ssh_url_text", ""))
        self.ssh_username_text = tk.StringVar()
        self.ssh_username_text.set(ConfigDict.get("ssh_username_text", ""))
        self.ssh_password_text = tk.StringVar()
        self.ssh_password_text.set(ConfigDict.get("ssh_password_text", ""))

        self.subscribeList = []
        self.caseList = []
        self.caseResList: [dict] = []

    def set_mqtt_connect(self):
        mqtt_connect = tk.LabelFrame(self.right_frame, text="连接",
                                     # width=200, height=200
                                     )
        tk.Label(mqtt_connect, text="服务器地址:").grid(row=0, column=0, sticky="w")
        tk.Label(mqtt_connect, text="连接状态:").grid(row=1, column=0, sticky="w")
        self.Entry_mqtt_connect = tk.Entry(mqtt_connect, textvariable=self.mqtt_url_text, width=30)
        self.Entry_mqtt_connect.grid(row=0, column=1, padx=4)

        self.Label_mqtt_state = tk.Label(mqtt_connect, text="未连接")
        self.Label_mqtt_state.grid(row=1, column=1, sticky=tk.W)
        self.btn_mqtt_connect = tk.Button(mqtt_connect, text="连接", width=8, command=self.click_mqtt_connect)
        self.btn_mqtt_connect.grid(row=0, column=2, padx=4, pady=2, rowspan=2)
        mqtt_connect.pack(anchor=tk.W)

    def set_ssh_connect(self):
        ssh_connect = tk.LabelFrame(self.right_frame, text="ssh连接",
                                    # width=200, height=200
                                    )
        tk.Label(ssh_connect, text="ssh地址:").grid(row=0, column=0, sticky="w")
        tk.Label(ssh_connect, text="用户名:").grid(row=1, column=0, sticky="w")
        tk.Label(ssh_connect, text="密码:").grid(row=2, column=0, sticky="w")

        Entry_width = 43

        self.Entry_ssh_url_text = tk.Entry(ssh_connect, textvariable=self.ssh_url_text, width=Entry_width)
        self.Entry_ssh_url_text.grid(row=0, column=1, padx=4, columnspan=2)
        self.Entry_ssh_username_text = tk.Entry(ssh_connect, textvariable=self.ssh_username_text, width=Entry_width)
        self.Entry_ssh_username_text.grid(row=1, column=1, padx=4, columnspan=2)
        self.Entry_ssh_password_text = tk.Entry(ssh_connect, textvariable=self.ssh_password_text, width=Entry_width)
        self.Entry_ssh_password_text.grid(row=2, column=1, padx=4, columnspan=2)
        ssh_connect.pack(anchor=tk.W)

    def set_subscribe(self):
        subscription = tk.LabelFrame(self.center_frame, text="订阅")
        self.subscription_state = tk.StringVar()
        self.subscription_state.set("")
        tk.Label(subscription, textvariable=self.subscription_state, width=8).grid(row=0, column=0, padx=2, pady=2,
                                                                                   sticky="e")
        tk.Button(subscription, text="导入", width=6, command=self.open_subscription).grid(row=0, column=2, padx=2,
                                                                                           pady=2, sticky="e")
        # tk.Button(subscription, text="导出", width=6).grid(row=0, column=2, padx=2, pady=2, sticky="e")
        tk.Button(subscription, text="订阅", width=6, command=self.start_subscription).grid(row=0, column=3, padx=2,
                                                                                            pady=2, sticky="e")
        # tk.Button(subscription, text="取消", width=6).grid(row=0, column=4, padx=2, pady=2, sticky="e")
        frame = tk.Frame(subscription)
        columns = ["序号", "主题", "请求"]
        ybar = ttk.Scrollbar(frame, orient='vertical')  # 垂直滚动条
        ybar.pack(side=tk.RIGHT, fill=tk.Y)
        self.TreeView_subscription = ttk.Treeview(
            master=frame,  # 父容器
            height=5,  # 表格显示的行数,height行
            columns=columns,  # 显示的列
            displaycolumns="#all",
            show='headings',  # 隐藏首列
            yscrollcommand=ybar.set,
        )
        ybar['command'] = self.TreeView_subscription.yview
        self.TreeView_subscription.heading('序号', text='序号')  # 定义表头
        self.TreeView_subscription.heading('主题', text='主题')  # 定义表头
        self.TreeView_subscription.heading('请求', text='请求')  # 定义表头
        # self.TreeView_subscription.column('#0', width=50, anchor=tk.S, )  # 定义列
        self.TreeView_subscription.column('序号', width=40, anchor=tk.S, )  # 定义列
        self.TreeView_subscription.column('主题', width=330, anchor=tk.S, )  # 定义列
        self.TreeView_subscription.column('请求', width=40, anchor=tk.S, )  # 定义列
        # self.TreeView_subscription.grid(row=1, column=0, sticky='e', columnspan=5)
        self.TreeView_subscription.pack()
        frame.grid(row=1, column=0, sticky='e', columnspan=5)
        subscription.pack(anchor=tk.W)

    def set_message_validate(self):
        frame = tk.LabelFrame(self.center_frame, text='报文解析')
        self.Text_message_validate = tk.Text(frame, width=10, height=10)
        self.Text_message_validate.pack(fill=tk.X)
        tk.Button(frame, text='解析', command=self.parse_message_validate).pack(fill=tk.X, pady=2, padx=2)
        frame.pack(anchor=tk.W, fill=tk.X, pady=6)

    def set_message_result(self):
        frame = tk.LabelFrame(self.left_frame, text='报文解析结果')
        self.Text_message_result = tk.Text(frame, width=60, height=27)
        self.Text_message_result.pack(fill=tk.X)
        frame.pack(anchor=tk.W, fill=tk.X)

    def set_case_table(self):
        table_frame = tk.LabelFrame(self.right_frame, text='案例')
        btn_frame = tk.Frame(table_frame)
        tk.Button(btn_frame, text="导入", width=6, command=self.open_case_file). \
            grid(row=0, column=0, padx=20, pady=2)
        tk.Button(btn_frame, text="执行", width=6, command=lambda: thread_it(self.start_case)). \
            grid(row=0, column=1, padx=20, pady=2)
        tk.Button(btn_frame, text="导出结果", width=6, command=self.export_case_result). \
            grid(row=0, column=2, padx=20, pady=2)
        btn_frame.pack(fill=tk.X)

        frame = tk.Frame(table_frame)
        columns = ["序号", "用例名称"]
        ybar = ttk.Scrollbar(frame, orient='vertical')  # 垂直滚动条
        ybar.pack(side=tk.RIGHT, fill=tk.Y)
        self.TreeView_case = ttk.Treeview(
            master=frame,  # 父容器
            height=25,  # 表格显示的行数,height行
            columns=columns,  # 显示的列
            displaycolumns="#all",
            show='headings',  # 隐藏首列
            yscrollcommand=ybar.set,
        )
        ybar['command'] = self.TreeView_case.yview
        self.TreeView_case.heading('序号', text='序号')  # 定义表头
        self.TreeView_case.heading('用例名称', text='用例名称')  # 定义表头
        # self.TreeView_case.column('#0', width=50, anchor=tk.S, )  # 定义列
        self.TreeView_case.column('序号', width=40, anchor=tk.S, )  # 定义列
        self.TreeView_case.column('用例名称', width=300, anchor=tk.S, )  # 定义列
        # self.TreeView_case.grid(row=1, column=0, sticky='e', columnspan=5)
        self.TreeView_case.pack()
        frame.pack(anchor=tk.W, fill=tk.X, side="bottom")
        table_frame.pack(anchor=tk.W, fill=tk.X)

    def set_case_result_table(self):
        table_frame = tk.LabelFrame(self.bottom_frame, text='案例结果')
        frame = tk.Frame(table_frame)
        columns = ["序号", "用例名称", "执行结果", "时间"]
        ybar = ttk.Scrollbar(frame, orient='vertical')  # 垂直滚动条
        ybar.pack(side=tk.RIGHT, fill=tk.Y)
        self.TreeView_case_result = ttk.Treeview(
            master=frame,  # 父容器
            height=15,  # 表格显示的行数,height行
            columns=columns,  # 显示的列
            displaycolumns="#all",
            show='headings',  # 隐藏首列
            yscrollcommand=ybar.set,
        )
        ybar['command'] = self.TreeView_case_result.yview
        self.TreeView_case_result.heading('序号', text='序号')  # 定义表头
        self.TreeView_case_result.heading('用例名称', text='用例名称')  # 定义表头
        self.TreeView_case_result.heading('执行结果', text='执行结果')  # 定义表头
        self.TreeView_case_result.heading('时间', text='时间')  # 定义表头
        # self.TreeView_case_result.column('#0', width=50, anchor=tk.S, )  # 定义列
        self.TreeView_case_result.column('序号', width=40, anchor=tk.S, )  # 定义列
        self.TreeView_case_result.column('用例名称', width=360, anchor=tk.S, )  # 定义列
        self.TreeView_case_result.column('执行结果', width=120, anchor=tk.S, )  # 定义列
        self.TreeView_case_result.column('时间', width=340, anchor=tk.S, )  # 定义列
        # self.TreeView_case_result.grid(row=1, column=0, sticky='e', columnspan=5)
        self.TreeView_case_result.pack()
        self.TreeView_case_result.bind("<<TreeviewSelect>>", self.select_case_result)
        frame.pack(fill=tk.X)
        table_frame.pack(anchor=tk.W, fill=tk.X)

    def parse_message_validate(self):
        self.Text_message_result.delete("1.0", tk.END)
        payload = self.Text_message_validate.get("1.0", tk.END)
        res = ParseMsg().parse_payloadHex(payload)
        self.Text_message_result.insert(tk.INSERT, self.transform_dict(res))

    def click_mqtt_connect(self):
        # print(self.mqtt_url_text.get())
        if self.mqtt_url_text.get().find(":") == -1:
            self.Label_mqtt_state["text"] = "格式错误！"
            return
        hostname = self.mqtt_url_text.get().split(':')[0]
        port = int(self.mqtt_url_text.get().split(':')[1])
        if mqtt_client.connect(hostname, port, 600):
            self.btn_mqtt_connect.config(text="断开连接", command=self.click_mqtt_disconnect)
            self.Entry_mqtt_connect.config(state=tk.DISABLED)
            self.Label_mqtt_state["text"] = "已连接"
        else:
            self.Label_mqtt_state["text"] = "连接失败"

    def click_mqtt_disconnect(self):
        mqtt_client.close_client()
        self.btn_mqtt_connect.config(text="连接", command=self.click_mqtt_connect)
        self.Entry_mqtt_connect.config(state=tk.NORMAL)
        self.Label_mqtt_state["text"] = "已断开连接"

    def click_ssh_disconnect(self):
        ssh_client.client.close()
        # self.Entry_ssh_url_text.config(state=tk.NORMAL)
        # self.Entry_ssh_username_text.config(state=tk.NORMAL)
        # self.Entry_ssh_password_text.config(state=tk.NORMAL)
        # self.Btn_ssh_connect.config(state=tk.NORMAL)
        # self.Label_ssh_state["text"] = "已断开连接"

    def open_subscription(self):
        filename = askopenfilename(filetypes=[('Excel表格', '*.xlsx')])
        if filename == "":
            return
        wb = openpyxl.load_workbook(filename)
        sheet = wb[wb.sheetnames[0]]
        self.subscribeList = []
        for row in range(sheet.max_row - 1):
            tmp = []
            tmp.append(row)
            tmp.append(sheet.cell(row + 2, 2).value)
            tmp.append(sheet.cell(row + 2, 3).value.split('-')[0])
            self.subscribeList.append(tmp)
        self.show_subscription()
        # print(self.subscribeList)

    def open_case_file(self):
        filename = askopenfilename(filetypes=[('Excel表格', '*.xlsx')])
        self.TreeView_case.delete(*self.TreeView_case.get_children())
        self.caseList = []
        if filename == "":
            return
        self.TreeView_case_result.delete(*self.TreeView_case_result.get_children())
        self.caseResList = []
        wb = openpyxl.load_workbook(filename)
        sheet = wb[wb.sheetnames[0]]
        # ['状态码', '用例名称', 'ssh_url', 'ssh_user', 'ssh_pwd', 'ssh_cmd', 'publish_topic', 'publish_mid']
        col_name = []

        for col in range(sheet.max_column):
            col_name.append(sheet.cell(1, col + 1).value)

        self.caseList = []
        for row in range(1, sheet.max_row):
            tmp = defaultdict(str)
            for col in range(sheet.max_column):
                if sheet.cell(row + 1, 1).value is None:
                    break
                tmp[col_name[col]] = sheet.cell(row + 1, col + 1).value
            if tmp == {}:
                break
            self.caseList.append(tmp)
        self.show_case()

    def show_case(self):
        self.TreeView_case.delete(*self.TreeView_case.get_children())
        for idx, item in enumerate(self.caseList):
            value = [idx, item["用例名称"]]
            self.TreeView_case.insert('', 'end', values=value)

    def show_subscription(self):
        self.TreeView_subscription.delete(*self.TreeView_subscription.get_children())
        for row in range(len(self.subscribeList)):
            self.TreeView_subscription.insert('', 'end', values=self.subscribeList[row])

    def start_case(self):
        if not self.caseList:
            print("案例列表为空")
            return

        if self.ssh_url_text.get().find(":") == -1:
            self.Text_message_result.delete("1.0", tk.END)
            self.Text_message_result.insert(tk.END, "ssh_url 格式错误!")
            return

        if self.btn_mqtt_connect["text"] != "断开连接":
            self.click_mqtt_connect()
            if self.btn_mqtt_connect["text"] != "断开连接":
                self.Text_message_result.delete("1.0", tk.END)
                self.Text_message_result.insert(tk.END, "mqtt连接失败")
                return

        if self.subscription_state.get() != "订阅成功":
            self.start_subscription()
            if self.subscription_state.get() != "订阅成功":
                self.Text_message_result.delete("1.0", tk.END)
                self.Text_message_result.insert(tk.END, "订阅失败")
                return

        self.ssh_info = {
            "hostname": self.ssh_url_text.get().split(':')[0],
            "port": int(self.ssh_url_text.get().split(':')[1]),
            "username": self.ssh_username_text.get(),
            "password": self.ssh_password_text.get(),
        }

        self.btn_mqtt_connect.config(state=tk.DISABLED)
        self.Entry_ssh_url_text.config(state=tk.DISABLED)
        self.Entry_ssh_username_text.config(state=tk.DISABLED)
        self.Entry_ssh_password_text.config(state=tk.DISABLED)

        self.caseResList = []

        # 清空结果列表
        self.TreeView_case_result.delete(*self.TreeView_case_result.get_children())
        self.window.update()
        PyLog().set_logFileName()
        for idx, case in enumerate(self.caseList):
            print(f"index: {idx}")
            PyLog().set_log(f"============================== 开始测试【序号:{idx}】【用例名称】:"
                            + case.get("用例名称") + "==============================")
            self.caseResList.append(self.start_case_single(case))
            self.insert_case_result(idx)
            PyLog().set_log(f'============================== 【{case.get("用例名称")}】测试结束 ==============================')

        # for idx, case in enumerate(self.caseList):
        #     self.caseResList.append(self.start_case_single(case))
        # self.show_case_result()

        self.btn_mqtt_connect.config(state=tk.NORMAL)
        self.Entry_ssh_url_text.config(state=tk.NORMAL)
        self.Entry_ssh_username_text.config(state=tk.NORMAL)
        self.Entry_ssh_password_text.config(state=tk.NORMAL)

    def start_case_single(self, case):
        py_case = PyCase()
        py_case.load_ssh_info(self.ssh_info)
        return py_case.get_res(case, mqtt_client, ssh_client)

    def start_subscription(self):
        if mqtt_client.isConnected is False:
            self.subscription_state.set("订阅失败")
            return
        subscription_list = [(item[1], int(item[2])) for item in self.subscribeList]
        # print(subscription_list)
        mqtt_client.client.subscribe(subscription_list)
        mqtt_client.loop_start()
        self.subscription_state.set("订阅成功")
        # print(mqtt_client.queue.get())

    def show_case_result(self):
        self.TreeView_case_result.delete(*self.TreeView_case_result.get_children())
        for idx, item in enumerate(self.caseResList):
            case_res = "成功" if item["is_Success"] else "失败"
            value = [idx, self.caseList[idx]["用例名称"], case_res, item["time"]]
            self.TreeView_case_result.insert('', 'end', values=value, iid=str(idx))

    def insert_case_result(self, idx: int):
        case_res = "成功" if self.caseResList[idx]["is_Success"] else "失败"
        if case_res == "失败":
            print(f"{self.caseResList[idx]}")
            print(self.caseResList[idx]["error_info"])
        value = [idx, self.caseList[idx]["用例名称"], case_res, self.caseResList[idx]["time"]]
        self.TreeView_case_result.insert('', 'end', values=value, iid=str(idx))
        self.window.update()

    def select_case_result(self, event):
        idx = int(self.TreeView_case_result.selection()[0])
        self.pop_window(idx)

    def export_case_result(self):
        if not self.caseResList:
            return

        data_head = time.strftime("%Y%m%d-%H%M%S", time.localtime())
        os.makedirs("res", exist_ok=True)
        filename = f"res/res_{data_head}.xlsx"

        col_names = ["序号", "用例名称", "执行结果", "时间"]
        col_names.extend(list(CaseResData.keys()))
        col_names.extend(["error_info", "message", "topic"])

        wb = openpyxl.Workbook()

        sheet = wb[wb.sheetnames[0]]

        for idx, name in enumerate(col_names):
            sheet.cell(1, idx + 1).value = name
        for idx in range(len(self.caseResList)):
            sheet.cell(idx + 2, 1).value = idx
            sheet.cell(idx + 2, 2).value = self.caseList[idx]["用例名称"]
            case_res = "成功" if self.caseResList[idx]["is_Success"] else "失败"
            sheet.cell(idx + 2, 3).value = case_res
            sheet.cell(idx + 2, 4).value = self.caseResList[idx]["time"]
            tmp_idx = 5
            for key in CaseResData.keys():
                sheet.cell(idx + 2, tmp_idx).value = str(self.caseResList[idx]["data"][key])
                tmp_idx += 1
            for key in col_names[tmp_idx - 1:]:
                value = self.caseResList[idx][key]
                if key == "message":
                    value = value.hex()
                sheet.cell(idx + 2, tmp_idx).value = str(value)
                tmp_idx += 1
        wb.save(filename)
        self.Text_message_result.delete("1.0", tk.END)
        self.Text_message_result.insert(tk.END, "导出结果执行完毕")

    def pop_window(self, idx):
        child_window = tk.Toplevel(self.window)
        child_window.title(str(idx) + "-" + self.caseList[idx]["用例名称"])
        frame = tk.LabelFrame(child_window, text='报文解析结果')
        self.Text_case_result = tk.Text(frame, width=60, height=27)
        self.Text_case_result.pack(fill=tk.X)
        frame.pack(anchor=tk.W, fill=tk.X)

        self.Text_case_result.insert(tk.END, self.transform_dict(self.caseResList[idx]["data"]))

        tkinter_tool.set_center(child_window, 700, 400)
        self.window.attributes('-disabled', True)
        self.window.wait_window(child_window)
        self.window.attributes('-disabled', False)
        self.window.deiconify()

    @staticmethod
    def transform_dict(d: dict):
        text = ""
        for key, value in d.items():
            if key == "断言结果":
                continue
            text += f"{key}: {value}\n"
        return text


if __name__ == '__main__':
    UI = tk_ui()

    tkinter_tool.set_center(UI.window, 1280, 750)
    UI.window.mainloop()
