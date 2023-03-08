import openpyxl


def get_cases(file_name, sheet):
    sheet = read_case(file_name, sheet)
    rows = sheet.max_row
    list_cases = []
    for row in range(2, rows + 1):
        dict_case = dict(
            id=sheet.cell(row, 1).value,
            url=sheet.cell(row, 2).value,
            start=sheet.cell(row, 3).value,
            end=sheet.cell(row, 4).value)
        if dict_case["start"] is None:
            break
        list_cases.append(dict_case)
    return list_cases


def read_case(file_name, sheet):
    return openpyxl.load_workbook(file_name)[sheet]


if __name__ == '__main__':
    res = get_cases('采集录入.xlsx', "Sheet1")
